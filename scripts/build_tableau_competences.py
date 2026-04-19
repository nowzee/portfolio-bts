"""
Génère le tableau de synthèse E5 BTS SIO 2026 pour Louis PERROUX
à partir du modèle officiel Annexe VI-1.
"""
import shutil
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = Path(r"C:/Users/sachu/Downloads/BTS_SIO_Annexe_6_Epreuve E5 - Tableau de synthe_se_2026 (1).xlsx")
OUT = ROOT / "assets" / "files" / "tableau-competences.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

# 1) on duplique le modèle pour préserver tout le formatage officiel
shutil.copy(TEMPLATE, OUT)

wb = load_workbook(OUT)
ws = wb.active

# ============ Identité ============
ws["A3"] = "NOM et prénom : PERROUX Louis"
ws["F3"] = "N° candidat : 02246416792"
ws["A4"] = "Centre de formation : ESNA de Bretagne - 6 Rue Maryse Bastié, 35170 Bruz"
ws["G4"] = "▢ SISR"
ws["H4"] = "X SLAM"
ws["A5"] = "Adresse URL du portfolio : https://nowzee.github.io/portfolio-bts"

# ============ Helpers ============
# Colonnes des compétences (exactement comme dans le modèle):
# C=Gérer le patrimoine    D=Incidents/assistance    E=Présence en ligne
# F=Mode projet            G=Mise à dispo service    H=Dév. professionnel
COMP_COLS = ("C", "D", "E", "F", "G", "H")

def clear_row(row: int):
    """Vide explicitement les colonnes A à M de la ligne donnée."""
    for col in range(1, 14):
        ws.cell(row=row, column=col).value = None

def fill_row(row: int, libelle: str, periode: str, comps: str):
    """comps = chaîne de 6 caractères (X ou .) pour C D E F G H."""
    assert len(comps) == 6
    clear_row(row)
    if libelle is not None:
        ws[f"A{row}"].value = libelle
    if periode is not None:
        ws[f"B{row}"].value = periode
    for i, ch in enumerate(comps):
        if ch.upper() == "X":
            ws[f"{COMP_COLS[i]}{row}"].value = "X"

# ============ Réalisations en cours de formation (rows 9-18) ============
# Ordre: C=patrimoine, D=incidents, E=présence en ligne, F=projet, G=service, H=dév pro

formation = [
    # libellé,                                                                   période,                   C D E F G H
    ("Threatlab — plateforme de honeypots (https://github.com/nowzee/Threatlab)",  "01/07/25 au 28/02/26", "XX..XX"),
    ("Yoda — outil forensic en Rust",                                              "01/09/25 au 15/04/26", "X..XXX"),
    ("RTERM — gestion de connexions SSH (Rust + Tauri, https://github.com/nowzee/RTerm)", "01/09/25 au 15/04/26", "...XXX"),
    ("Alice_In_Rans0ml4nd — challenge maker forensic ECW 2025",                    "01/03/25 au 30/04/25", "X..X.X"),
    ("Gh0st_1n_7h3_G1t — challenge maker MIDNIGHT FLAG 2026",                      "01/01/26 au 28/02/26", "...X.X"),
    ("Portfolio BTS (https://nowzee.github.io/portfolio-bts)",                     "01/01/26 au 15/04/26", "..XX.X"),
    ("Projet TX — projet de groupe ESNA",                                          "29/05/25 au 30/08/25", "XX.XXX"),
    ("Projet RADAR — outil de scan Active Directory (groupe ESNA)",                "10/11/25 au 20/11/25", "XX.XXX"),
    ("Veille technologique cybersécurité (page veille du portfolio)",              "01/09/24 au 15/04/26", "...X.X"),
]

start_row = 9
for i, (lib, per, comps) in enumerate(formation):
    fill_row(start_row + i, lib, per, comps)

# Vider d'éventuelles lignes restantes (10-18 max selon modèle)
for r in range(start_row + len(formation), 19):
    fill_row(r, None, None, "......")

# ============ Réalisations PILLET-HITECH 1ère année (rows 20-26) ============
pro_y1 = [
    ("Portail captif d'entreprise — PILLET-HITECH (développé seul, revendu aux clients)",
     "01/09/24 au 31/08/25", "XXXXXX"),
]
start_row = 20
for i, (lib, per, comps) in enumerate(pro_y1):
    fill_row(start_row + i, lib, per, comps)
for r in range(start_row + len(pro_y1), 27):
    fill_row(r, None, None, "......")

# ============ Réalisations PILLET-HITECH 2ème année (rows 28-34) ============
pro_y2 = [
    ("Sauvegarde automatique Proxmox / Zimbra avec rsync — PILLET-HITECH",
     "01/09/25 au 31/12/25", "XX.XXX"),
]
start_row = 28
for i, (lib, per, comps) in enumerate(pro_y2):
    fill_row(start_row + i, lib, per, comps)
for r in range(start_row + len(pro_y2), 35):
    fill_row(r, None, None, "......")

wb.save(OUT)
print(f"OK -> {OUT}")
print(f"  Formation : {len(formation)} réalisations")
print(f"  Pro Y1    : {len(pro_y1)} réalisation(s)")
print(f"  Pro Y2    : {len(pro_y2)} réalisation(s)")
